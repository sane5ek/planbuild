import os
import openpyxl as xlsx
from openpyxl.utils.cell import column_index_from_string
import pyexcel as p

import xlrd as xls

def handle_file(f):
    with open('name.jpg', 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)

def get_excel_subjects(path):
    path = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '\\' + path.replace('/', '\\')
    response = []
    if path[-1] == 's':
        # xls to xlsx
        p.save_book_as(file_name=path,
                       dest_file_name=path+'x')
        path += 'x'
    if path[-1] == 'x':
        # xlsx
        wb = xlsx.load_workbook(path, data_only=True)
        ws = wb.active
        subjects_starts = 0
        for i in range(1, 15):
            if ws['A'+str(i)].value == 'Дисциплина':
                for j in range(i+1, 20):
                    if ws['A'+str(j)].value != None:
                        subjects_starts = j
                        break
        semester = 0 # TODO: counting semester in ZAO
        course = 0
        for i in ws:
            for j in i:
                if j.value == 'Курс':
                    course = j.column
                if j.value == 'Семестр':
                    semester = j.column
        # если ячейка строковая и без цвета - это предмет
        for i in ws.iter_rows(min_row=subjects_starts, min_col=1, max_col=1):
            if i[0].data_type == 's' and not isinstance(i[0].font.color.rgb, str):
                current_subj = {}
                current_subj['number'] = len(response) + 1
                current_subj['subject'] = i[0].value
                if semester != 0:
                    current_subj['semester'] = list(ws.iter_cols(min_col=column_index_from_string(semester), min_row=i[0].row,
                                       max_col=column_index_from_string(semester), max_row=i[0].row))[0][0].value
                else:
                    current_subj['semester'] = 0
                if course != 0:
                    current_subj['course'] = list(ws.iter_cols(min_col=column_index_from_string(course), min_row=i[0].row,
                                       max_col=column_index_from_string(course), max_row=i[0].row))[0][0].value
                else:
                    current_subj['course'] = 0

                response.append(current_subj)
    else:
        raise ValueError('Not an Excel file')
    return response