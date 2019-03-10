import os
from copy import copy

import openpyxl as xlsx
import pyexcel as p
from django.core.files import File
from django.db import models
from django.utils import timezone
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

from .utils import get_absolute_path


# Create your models here.

class RequestType(models.Model):
    """Model representing a types of requests (e.g. to share your fields)"""
    name = models.CharField(max_length=100)

    def __str__(self):
        """String for representing the Model object (in Admin site etc.)"""
        return self.name


class Request(models.Model):
    sender = models.ForeignKey('builder_auth.CustomUser', related_name='Request_sender', null=False, default=0,
                               on_delete=models.CASCADE)
    receiver = models.ForeignKey('builder_auth.CustomUser', related_name='Request_receiver', null=False, default=0,
                                 on_delete=models.CASCADE)
    type = models.ForeignKey('builder.RequestType', null=False, default=0, on_delete=models.CASCADE)
    create_date = models.DateTimeField('Date of setting sender request', null=False, default=timezone.now())
    answer_date = models.DateTimeField('Date of getting receiver answer', null=True, default=None)
    result = models.ForeignKey('builder.RequestResultType', null=True, default=None, on_delete=models.SET_NULL)

    def __str__(self):
        return "From {0} to {1} type {2}".format(self.sender, self.receiver, self.type)


class RequestResultType(models.Model):
    name = models.CharField(max_length=10)

    def __str__(self):
        return self.name


class Load(models.Model):
    """Model representing a types of loads (e.g. 1_1M)"""
    name = models.CharField(max_length=10)

    def __str__(self):
        return self.name


class Post(models.Model):
    """Model representing of a types posts (e.g. Teacher)"""
    name = models.CharField(max_length=100)

    def __str__(self):
        return self.name


class ScienceDegree(models.Model):
    """Model representing a types of science degrees (e.g. )"""
    name = models.CharField(max_length=100)

    def __str__(self):
        return self.name


class ScienceTitle(models.Model):
    """Model representing a types of science titles (e.g. )"""
    name = models.CharField(max_length=100)

    def __str__(self):
        return self.name


class FieldManager(models.Manager):

    def FillTemplate(self, subjects, courses, diplomas, load_filename, template_filename, user):

        # print(template_filename)
        destiny_filename = 'Plan_{0}_{1}.xlsm'.format(user.id, str(timezone.now().date()))

        # receive type of load

        type_of_load_name = ''

        if '1_1' in load_filename:
            if 'ЗАО' in load_filename or 'ZAO' in load_filename:
                type_of_load_name = '1_1MZAO'
            else:
                type_of_load_name = '1_1M'
        elif '2_4' in load_filename:
            if 'ЗАО' in load_filename or 'ZAO' in load_filename:
                type_of_load_name = '2_4ZAO'
            else:
                type_of_load_name = '2_4'

        if type_of_load_name:
            type_of_load_id = Load.objects.filter(name=type_of_load_name)[0].id
        else:
            raise ValueError('Can\'t get type of the load document {0}'.format(load_filename))

        # receive all fields for that load for that owner

        fields = Field.objects.filter(owner=user.adopted_fields).filter(type_of_load_id=type_of_load_id)

        # НАЙТИ ВСЕ СТРОЧКИ ДЛЯ ВСТАВКИ (ДНЕВНОЙ БЮДЖЕТ И ТД)
        first_semester = {
            'day_budget': 5,
            'total_day_budget': 8,
            'day_contract': 9,
            'total_day_contract': 12,
            'total_day': 13,

            'night_budget': 14,
            'total_night_budget': 17,
            'night_contract': 18,
            'total_night_contract': 21,
            'total_night': 22,

            'semester_budget': 23,
            'semester_contract': 24,
            'total_semester': 25
        }

        second_semester = {
            'day_budget': 5,
            'total_day_budget': 8,
            'day_contract': 9,
            'total_day_contract': 12,
            'total_day': 13,

            'night_budget': 14,
            'total_night_budget': 17,
            'night_contract': 18,
            'total_night_contract': 21,
            'total_night': 22,

            'semester_budget': 23,
            'semester_contract': 24,
            'total_semester': 25,

            'total_year_budget': 26,
            'total_year_contract': 27,
            'total_year': 28
        }

        # get worksheets
        plan_workbook = xlsx.load_workbook(template_filename, keep_vba=True)
        load_worksheet = xlsx.load_workbook(load_filename, data_only=True).worksheets[0]
        first_semester_load_worksheet = plan_workbook.worksheets[1]
        second_semester_load_worksheet = plan_workbook.worksheets[2]

        # какие столбцы с формулами изменить
        start = fields.filter(name_in_load='Лекций')[0].column_in_plan
        # end = fields.filter(name_in_plan='ВСЕГО')[0].column_in_plan
        end = 24

        # УЗНАТЬ СТРОКИ ДЛЯ ВСТАВКИ ПРЕДМЕТОВ
        for subject in subjects:
            rows_to_insert = [0, 0]
            load_row = int(subject['#'])
            selection = fields.filter(name_in_load='Контингент')

            # количество студентов бюджета и договора
            budget = int(load_worksheet.cell(load_row, selection[0].column_in_load).value)
            contract = int(load_worksheet.cell(load_row, selection[1].column_in_load).value)

            # выбираем строки для вставки

            if int(subject['Семестр']) % 2 == 0:
                # второй семестр
                semester = second_semester
                worksheet = second_semester_load_worksheet
            else:
                # первый семестр
                semester = first_semester
                worksheet = first_semester_load_worksheet

            load_type = 0
            if budget != 0:
                load_type = 0
                # если бюджет и дневное
                if type_of_load_id <= 2:
                    if int(subject['Семестр']) % 2 == 0:
                        # второй семестр
                        rows_to_insert[0] = second_semester['total_day_budget'] - 1
                    else:
                        # первый семестр
                        rows_to_insert[0] = first_semester['total_day_budget'] - 1
                # если бюджет и заочное
                else:
                    if int(subject['Семестр']) % 2 == 0:
                        # второй семестр
                        rows_to_insert[0] = second_semester['total_night_budget'] - 1
                    else:
                        # первый семестр
                        rows_to_insert[0] = first_semester['total_night_budget'] - 1

            # инкрементируем все поля
            if rows_to_insert[0] != 0:
                for key in semester:
                    if semester[key] > rows_to_insert[0]:
                        semester[key] += 1

            if contract != 0:
                load_type = 1
                # если договор и дневное
                if type_of_load_id <= 2:
                    if int(subject['Семестр']) % 2 == 0:
                        # второй семестр
                        rows_to_insert[1] = second_semester['total_day_contract'] - 1
                    else:
                        # первый семестр
                        rows_to_insert[1] = first_semester['total_day_contract'] - 1
                    # если бюджет и заочное
                else:
                    if int(subject['Семестр']) % 2 == 0:
                        # второй семестр
                        rows_to_insert[1] = second_semester['total_night_contract'] - 1
                    else:
                        # первый семестр
                        rows_to_insert[1] = first_semester['total_night_contract'] - 1

            # инкрементируем все поля
            if rows_to_insert[1] != 0:
                for key in semester:
                    if semester[key] > rows_to_insert[1]:
                        semester[key] += 1

            if rows_to_insert[0] != 0:
                self._fill_subject_row(fields, load_worksheet, worksheet, load_row, rows_to_insert[0], 0,
                                       subject)
            if rows_to_insert[1] != 0:
                self._fill_subject_row(fields, load_worksheet, worksheet, load_row, rows_to_insert[1], 1,
                                       subject)

            # ЗАМЕНИТЬ ВСЕ ФОРМУЛЫ НА ТЕ, КОТОРЫЕ БУДУТ В РЕЗУЛЬТАТЕ ВСТАВКИ
            self._change_formulas(first_semester, second_semester, first_semester_load_worksheet,
                                  second_semester_load_worksheet, start, end)


        # заполнить курсовые и дипломные работы
        for course_work in courses:
            rows_to_insert = [0, 0]
            course = int(course_work['Курс'])
            semester = int(course_work['Семестр'])
            count = int(course_work['Количество'])
            # выбираем куда вставлять
            if course != 0 and count != 0:
                if semester % 2 == 0:
                    # второй семестр
                    rows_to_insert[1] = second_semester['total_day_budget'] - 1
                else:
                    # первый семестр
                    rows_to_insert[0] = first_semester['total_day_budget'] - 1

                # инкрементируем все поля
                semesters = [first_semester, second_semester]
                print(rows_to_insert)
                for i, row in enumerate(rows_to_insert, 0):
                    if row != 0:
                        for key in semesters[i]:
                            if semesters[i][key] > row:
                                semesters[i][key] += 1

                # вставляем

                self._insert_additional(course, count, rows_to_insert, fields, first_semester_load_worksheet,
                                        second_semester_load_worksheet, 'курсовые работы', 3, 'Курсовая работа')

                self._change_formulas(first_semester, second_semester, first_semester_load_worksheet,
                                      second_semester_load_worksheet, start, end)

        for diploma in diplomas:
            rows_to_insert = [0, 0]
            course = int(diploma['Курс'])
            semester = int(diploma['Семестр'])
            count = int(diploma['Количество'])
            # выбираем куда вставлять
            if course != 0 and count != 0:
                if semester % 2 == 0:
                    # второй семестр
                    rows_to_insert[1] = second_semester['total_day_budget'] - 1
                else:
                    # первый семестр
                    rows_to_insert[0] = first_semester['total_day_budget'] - 1

                # инкрементируем все поля
                semesters = [first_semester, second_semester]
                for i, row in enumerate(rows_to_insert, 0):
                    if row != 0:
                        for key in semesters[i]:
                            if semesters[i][key] > row:
                                semesters[i][key] += 1
                # вставляем
                self._insert_additional(course, count, rows_to_insert, fields, first_semester_load_worksheet,
                                        second_semester_load_worksheet, 'ВКР бакалавров', 20, 'Руководство ВКР')

                self._change_formulas(first_semester, second_semester, first_semester_load_worksheet,
                                      second_semester_load_worksheet, start, end)

        plan_workbook.save(destiny_filename)

        new_file = PlanFile(file=File(open(destiny_filename, 'rb')), owner=user)
        new_file.save()

        if os.path.exists(destiny_filename):
            os.remove(destiny_filename)
        else:
            print("The file does not exist")

    def _find_cell_column_with_value(self, worksheet, cell_value):
        for row in worksheet.iter_rows(min_col=1, min_row=1):
            for cell in row:
                try:
                    if cell.value == cell_value:
                        return cell.column
                except:
                    continue

    def _insert_additional(self, course, count, rows_to_insert, fields, first_semester_worksheet,
                           second_semester_worksheet, hours_filter, hours_multi, type):
        course_col = fields.filter(name_in_plan__contains='курс').first().column_in_plan
        count_col = fields.filter(name_in_plan__contains='студент').first().column_in_plan

        worksheet = first_semester_worksheet
        row = rows_to_insert[0]
        if rows_to_insert[1] != 0:
            worksheet = second_semester_worksheet
            row = rows_to_insert[1]
        self._insert_row_with_style(worksheet, row)

        hours_col = self._find_cell_column_with_value(worksheet, hours_filter)

        if not hours_col is None:
            worksheet.cell(row, 1).value = type
            worksheet.cell(row, course_col).value = course
            worksheet.cell(row, count_col).value = count
            worksheet.cell(row, hours_col).value = count * hours_multi

    def _change_formulas(self, first_semester, second_semester, first_worksheet, second_worksheet, start, end):

        for row in first_worksheet.iter_rows(min_row=first_semester['total_day_budget'], min_col=start,
                                             max_row=first_semester['total_day_budget'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '=SUM({0}{1}:{0}{2})'.format(get_column_letter(cell.column),
                                                              first_semester['day_budget'] + 1,
                                                              first_semester['total_day_budget'] - 1)
                except:
                    pass

        for row in first_worksheet.iter_rows(min_row=first_semester['total_day_contract'], min_col=start,
                                             max_row=first_semester['total_day_contract'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '=SUM({0}{1}:{0}{2})'.format(get_column_letter(cell.column),
                                                              first_semester['day_contract'] + 1,
                                                              first_semester['total_day_contract'] - 1)
                except:
                    pass

        for row in first_worksheet.iter_rows(min_row=first_semester['total_night_budget'], min_col=start,
                                             max_row=first_semester['total_night_budget'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '=SUM({0}{1}:{0}{2})'.format(get_column_letter(cell.column),
                                                              first_semester['night_budget'] + 1,
                                                              first_semester['total_night_budget'] - 1)
                except:
                    pass

        for row in first_worksheet.iter_rows(min_row=first_semester['total_night_contract'], min_col=start,
                                             max_row=first_semester['total_night_contract'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '=SUM({0}{1}:{0}{2})'.format(get_column_letter(cell.column),
                                                              first_semester['night_contract'] + 1,
                                                              first_semester['total_night_contract'] - 1)
                except:
                    pass

        for row in first_worksheet.iter_rows(min_row=first_semester['total_day'], min_col=start,
                                             max_row=first_semester['total_day'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '={0}{1}+{0}{2}'.format(get_column_letter(cell.column),
                                                         first_semester['total_day_contract'],
                                                         first_semester['total_day_budget'])
                except:
                    pass

        for row in first_worksheet.iter_rows(min_row=first_semester['total_night'], min_col=start,
                                             max_row=first_semester['total_night'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '={0}{1}+{0}{2}'.format(get_column_letter(cell.column),
                                                         first_semester['total_night_contract'],
                                                         first_semester['total_night_budget'])
                except:
                    pass

        for row in first_worksheet.iter_rows(min_row=first_semester['semester_budget'], min_col=start,
                                             max_row=first_semester['semester_budget'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '={0}{1}+{0}{2}'.format(get_column_letter(cell.column),
                                                         first_semester['total_night_budget'],
                                                         first_semester['total_day_budget'])
                except:
                    pass

        for row in first_worksheet.iter_rows(min_row=first_semester['semester_contract'], min_col=start,
                                             max_row=first_semester['semester_contract'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '={0}{1}+{0}{2}'.format(get_column_letter(cell.column),
                                                         first_semester['total_night_contract'],
                                                         first_semester['total_day_contract'])
                except:
                    pass

        for row in first_worksheet.iter_rows(min_row=first_semester['total_semester'], min_col=start,
                                             max_row=first_semester['total_semester'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '={0}{1}+{0}{2}'.format(get_column_letter(cell.column),
                                                         first_semester['semester_budget'],
                                                         first_semester['semester_contract'])
                except:
                    pass

        # -------------------------------------------------------------> ГОВНОКООООООД

        for row in second_worksheet.iter_rows(min_row=second_semester['total_day_budget'], min_col=start,
                                              max_row=second_semester['total_day_budget'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '=SUM({0}{1}:{0}{2})'.format(get_column_letter(cell.column),
                                                              second_semester['day_budget'] + 1,
                                                              second_semester['total_day_budget'] - 1)
                except:
                    pass

        for row in second_worksheet.iter_rows(min_row=second_semester['total_day_contract'], min_col=start,
                                              max_row=second_semester['total_day_contract'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '=SUM({0}{1}:{0}{2})'.format(get_column_letter(cell.column),
                                                              second_semester['day_contract'] + 1,
                                                              second_semester['total_day_contract'] - 1)
                except:
                    pass

        for row in second_worksheet.iter_rows(min_row=second_semester['total_night_budget'], min_col=start,
                                              max_row=second_semester['total_night_budget'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '=SUM({0}{1}:{0}{2})'.format(get_column_letter(cell.column),
                                                              second_semester['night_budget'] + 1,
                                                              second_semester['total_night_budget'] - 1)
                except:
                    pass

        for row in second_worksheet.iter_rows(min_row=second_semester['total_night_contract'], min_col=start,
                                              max_row=second_semester['total_night_contract'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '=SUM({0}{1}:{0}{2})'.format(get_column_letter(cell.column),
                                                              second_semester['night_contract'] + 1,
                                                              second_semester['total_night_contract'] - 1)
                except:
                    pass

        for row in second_worksheet.iter_rows(min_row=second_semester['total_day'], min_col=start,
                                              max_row=second_semester['total_day'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '={0}{1}+{0}{2}'.format(get_column_letter(cell.column),
                                                         second_semester['total_day_contract'],
                                                         second_semester['total_day_budget'])
                except:
                    pass

        for row in second_worksheet.iter_rows(min_row=second_semester['total_night'], min_col=start,
                                              max_row=second_semester['total_night'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '={0}{1}+{0}{2}'.format(get_column_letter(cell.column),
                                                         second_semester['total_night_contract'],
                                                         second_semester['total_night_budget'])
                except:
                    pass

        for row in second_worksheet.iter_rows(min_row=second_semester['semester_budget'], min_col=start,
                                              max_row=second_semester['semester_budget'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '={0}{1}+{0}{2}'.format(get_column_letter(cell.column),
                                                         second_semester['total_night_budget'],
                                                         second_semester['total_day_budget'])
                except:
                    pass

        for row in second_worksheet.iter_rows(min_row=second_semester['semester_contract'], min_col=start,
                                              max_row=second_semester['semester_contract'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '={0}{1}+{0}{2}'.format(get_column_letter(cell.column),
                                                         second_semester['total_night_contract'],
                                                         second_semester['total_day_contract'])
                except:
                    pass

        for row in second_worksheet.iter_rows(min_row=second_semester['total_semester'], min_col=start,
                                              max_row=second_semester['total_semester'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '={0}{1}+{0}{2}'.format(get_column_letter(cell.column),
                                                         second_semester['semester_budget'],
                                                         second_semester['semester_contract'])
                except:
                    pass

        # ТОТАЛЫ
        for row in second_worksheet.iter_rows(min_row=second_semester['total_year_budget'], min_col=start,
                                              max_row=second_semester['total_year_budget'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '={0}{1}+\'I1\'!{0}{2}'.format(get_column_letter(cell.column),
                                                                second_semester['semester_budget'],
                                                                first_semester['semester_budget'])
                except:
                    pass

        for row in second_worksheet.iter_rows(min_row=second_semester['total_year_contract'], min_col=start,
                                              max_row=second_semester['total_year_contract'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '={0}{1}+\'I1\'!{0}{2}'.format(get_column_letter(cell.column),
                                                                second_semester['semester_contract'],
                                                                first_semester['semester_contract'])
                except:
                    pass

        for row in second_worksheet.iter_rows(min_row=second_semester['total_year'], min_col=start,
                                              max_row=second_semester['total_year'], max_col=end):
            for cell in row:
                try:
                    cell.data_type = 'f'
                    cell.value = '={0}{1}+{0}{2}'.format(get_column_letter(cell.column),
                                                         second_semester['total_year_contract'],
                                                         second_semester['total_year_budget'])
                except:
                    pass

    def _insert_row_with_style(self, worksheet, row):
        # вставляем строку
        worksheet.insert_rows(row, 1)

        # копируем стили
        for i, cell in enumerate(
                worksheet.iter_cols(min_col=1, min_row=row + 1, max_row=row + 1),
                1):
            copy_cell = worksheet.cell(row, i)
            if cell[0].data_type == 'f':
                trans = Translator(cell[0].value, str(cell[0].column_letter) + str(cell[0].row - 1))
                copy_cell.value = str(
                    trans.translate_formula(str(copy_cell.column_letter) + str(copy_cell.row)))
                cell[0].value = str(
                    trans.translate_formula(str(cell[0].column_letter) + str(cell[0].row)))
            copy_cell.font = copy(cell[0].font)
            copy_cell.border = copy(cell[0].border)
            copy_cell.fill = copy(cell[0].fill)
            copy_cell.number_format = copy(cell[0].number_format)
            copy_cell.protection = copy(cell[0].protection)
            copy_cell.alignment = copy(cell[0].alignment)

    def _fill_subject_row(self, fields, load_worksheet, plan_worksheet, load_row, plan_row, load_type, subject):
        if load_worksheet.cell(load_row, fields.filter(load_type=load_type)[1].column_in_load).value != 0:

            self._insert_row_with_style(plan_worksheet, plan_row)

            # вставляем данные
            plan_worksheet.cell(plan_row, 1).value = subject['Предмет']
            # load_type: 0 - бюджет, 1 - контракт
            for field in fields.filter(load_type=load_type):
                load_column = field.column_in_load
                plan_column = field.column_in_plan
                if load_column != 0:
                    plan_worksheet.cell(plan_row, plan_column).value = load_worksheet.cell(load_row, load_column).value


class Field(models.Model):
    column_in_load = models.PositiveIntegerField('Column in load')
    column_in_plan = models.PositiveIntegerField('Column in plan')
    name_in_load = models.CharField('Name in load', max_length=100)
    name_in_plan = models.CharField('Name in plan', max_length=100)
    type_of_load = models.ForeignKey('Load', null=False, default=1, on_delete=models.CASCADE)
    owner = models.ForeignKey('builder_auth.CustomUser', null=True, default=None, on_delete=models.SET_NULL)
    load_type = models.BooleanField('Load type', default=False)

    objects = FieldManager()

    def __str__(self):
        return 'Field {0} for {1}'.format(self.name_in_load, self.type_of_load.name)


class Cell(models.Model):
    column = models.PositiveIntegerField('Column')
    row = models.PositiveIntegerField('Row')
    column_span = models.PositiveIntegerField('Column span')
    row_span = models.PositiveIntegerField('Row span')
    advanced = models.CharField('Advanced', max_length=3)
    font_size = models.PositiveSmallIntegerField('Font size')
    direction = models.BooleanField('Direction')

    VERTICAL_ALIGN = (
        ('t', 'Top'),
        ('m', 'Middle'),
        ('b', 'Bottom'),
    )

    HORIZONTAL_ALIGN = (
        ('l', 'Left'),
        ('c', 'Center'),
        ('r', 'Right')
    )

    vertical_align = models.CharField('Vertical align', max_length=1, choices=VERTICAL_ALIGN, blank=True,
                                      default='m')
    horizontal_align = models.CharField('Horizontal align', max_length=1, choices=HORIZONTAL_ALIGN, blank=True,
                                        default='h')
    width = models.PositiveIntegerField('Width')
    height = models.PositiveIntegerField('Height')
    value = models.CharField('Value', max_length=500)
    owner = models.ForeignKey('builder_auth.CustomUser', null=True, default=None, on_delete=models.SET_NULL)

    def __str__(self):
        return 'Cell {0}:{1} for template'.format(self.column, self.row)


class UploadFile(models.Model):
    file = models.FileField(upload_to='files/loads/%Y/%m/%d')
    owner = models.ForeignKey('builder_auth.CustomUser', null=True, default=None, on_delete=models.SET_NULL)

    def convert_file_to_xlsx(self):
        if self.file.name.endswith('xls'):
            # xls to xlsx
            p.save_book_as(file_name=self.file.name,
                           dest_file_name=self.file.name + 'x')
            self.file.name += 'x'
            self.save()

        elif not str(self.file).endswith('xlsx'):
            raise ValueError('Not Excel file')


class PlanFile(models.Model):
    file = models.FileField(upload_to='files/plans/%Y/%m/%d')
    owner = models.ForeignKey('builder_auth.CustomUser', null=True, default=None, on_delete=models.SET_NULL)


class SubjectManager(models.Manager):
    def _get_subjects_start_column(self, worksheet):
        for i in range(1, 15):
            if worksheet['A' + str(i)].value == 'Дисциплина':
                for j in range(i + 1, 20):
                    if worksheet['A' + str(j)].value != None:
                        return j

    def _get_course_and_semester(self, worksheet):
        course = 0
        semester = 0
        for row in worksheet.iter_rows(min_col=1, min_row=1):
            for col in row:
                try:
                    if col.value == 'Курс':
                        course = col.column
                    if col.value == 'Семестр':
                        semester = col.column
                except:
                    continue
        return course, semester

    def set_objects_from_excel(self, path, owner):

        # delete old subjects
        self.filter(owner=owner).delete()

        path = get_absolute_path(path)
        wb = xlsx.load_workbook(path, data_only=True)
        ws = wb.active

        subjects_starts = self._get_subjects_start_column(ws)
        # TODO: counting semester in ZAO
        course, semester = self._get_course_and_semester(ws)

        # если ячейка строковая и без цвета - это предмет
        subjects = []
        for i in ws.iter_rows(min_row=subjects_starts, min_col=1, max_col=1):
            if i[0].data_type == 's' and not isinstance(i[0].font.color.rgb, str):
                current_subj = Subject()
                current_subj.row_number = i[0].row
                current_subj.name = i[0].value
                if semester != 0:
                    current_subj.semester = ws.cell(column=semester, row=i[0].row).value
                    if current_subj.semester is None:
                        current_subj.semester = 0
                else:
                    current_subj.semester = 0
                if course != 0:
                    current_subj.course = ws.cell(column=course, row=i[0].row).value
                    if current_subj.course is None:
                        current_subj.course = 0
                else:
                    current_subj.course = 0
                current_subj.owner = owner
                subjects.append(current_subj)
        self.bulk_create(subjects)

    def get_objects_json(self, owner):
        subjects = self.filter(owner=owner)
        json = []
        for subject in subjects:
            subj = {}
            subj['number'] = subject.row_number
            subj['subject'] = subject.name
            subj['semester'] = subject.semester
            subj['course'] = subject.course
            json.append(subj)
        return json


class Subject(models.Model):
    semester = models.CharField(max_length=3, blank=False)
    course = models.CharField(max_length=3, blank=False)
    name = models.CharField(max_length=50, blank=False)
    row_number = models.IntegerField(null=False)
    owner = models.ForeignKey('builder_auth.CustomUser', null=True, default=None, on_delete=models.SET_NULL)

    objects = SubjectManager()

    def __str__(self):
        return 'Subject {0} for {1} course {2} semester'.format(self.name, self.course, self.semester)
